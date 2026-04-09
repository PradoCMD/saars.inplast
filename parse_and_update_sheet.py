import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import os
import glob
import re
from datetime import datetime, timedelta
import pdfplumber

def ct(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()

def parse_num(v):
    if not v:
        return 0.0
    s = ct(v).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_romaneio_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    
    fname = os.path.basename(pdf_path)
    
    ordem_m = re.search(r"ORDEM\s+DE\s+CARGA[:\s]+(\d+)", text, re.I)
    ordem = ordem_m.group(1) if ordem_m else ""
    if not ordem:
        fn_m = re.search(r"ROMANEIO\s+(\d+)", fname, re.I)
        if fn_m:
            ordem = fn_m.group(1)

    pedidos = []
    ped_m = re.search(
        r"N°\s*UNICO.*?Valor\s*Total\s*\n(.*?)(?=MONTANTE)", text, re.S | re.I
    )
    if ped_m:
        for line in ped_m.group(1).strip().split("\n"):
            l = ct(line)
            if not l:
                continue
            m = re.match(r"^(\d+)\s+(\d+)\s+(\d+)\s*-\s*(.+?)\s+([\d.,]+)$", l)
            if m:
                pedidos.append({
                    "cidade": "", 
                    "parceiro": ct(m.group(4))
                })

    itens = {}
    UNITS = r"(?:PC|UN|PT|KG|CX|MT|M2|LT|FD|SC|RL|CJ|JG|BD|GL|TB|PR|DZ|CT|MH|MO|MM|PAR|PÇ|PE|M)"
    itens_m = re.search(
        r"PRODUTO\s+UND\.\s+QTD\s+NEG\.\s+QTD\s+VOL\.\s*\n(.*?)(?=TOTAL\s*:|\Z)",
        text,
        re.S | re.I,
    )
    if itens_m:
        for line in itens_m.group(1).strip().split("\n"):
            l = ct(line)
            if not l:
                continue
            m = re.match(
                rf"^(\d+)\s*-\s*(.+?)\s+({UNITS})\s+([\d.,]+)(?:\s+([\d.,]+))?\s*$",
                l,
                re.I,
            )
            if m:
                sku = ct(m.group(1))
                qty = parse_num(m.group(4))
                itens[sku] = itens.get(sku, 0) + qty

    cidade_parceiro = pedidos[0]["parceiro"] if pedidos else "Desconhecido"
    
    return {
        "ordem": ordem,
        "name": f"RM {ordem} - {cidade_parceiro}",
        "itens": itens,
        "file": fname
    }

def main():
    print("Iniciando processamento...")
    folder = "/Volumes/logistica/Automação Romaneio Logistica x PCP/"
    pdfs = sorted(glob.glob(os.path.join(folder, "ROMANEIO *.pdf")))
    pdfs = [p for p in pdfs if "NOTA" not in os.path.basename(p).upper()]
    
    print(f"Lendo {len(pdfs)} PDFs...")
    pdf_data = {}
    for p in pdfs:
        try:
            data = parse_romaneio_pdf(p)
            if data["ordem"]:
                pdf_data[data["ordem"]] = data
        except Exception as e:
            print(f"Erro ao ler {p}: {e}")

    print("Carregando planilha...")
    wb = openpyxl.load_workbook('./PEDIDOS TRANSPORTADORAS 2025 (1).xlsx', data_only=False)
    ws = wb['PROGRAMAÇÃO']

    # Read products mapping
    product_rows = {}
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str) and " - " in val:
            sku = val.split(" - ")[0].strip()
            product_rows[sku] = r

    # Read existing romaneios
    existing_romaneios = {}
    
    # Process pairs of columns starting from C (col 3)
    # C is Romaneio, D is Estoque INT. Or they can be ANY col, let's trace from 3 to max_col
    romaneio_groups = [] # list of dicts: {'date': datetime, 'col_data': int, 'col_formula': int, 'name': str, 'ordem': str}
    
    max_date = datetime(2023, 1, 1)
    
    col = 3
    while col <= ws.max_column:
        name = ws.cell(row=2, column=col).value
        
        if not name or not str(name).startswith("RM "):
            # Could be "ESTOQUE INT" or empty, skip to find real RM columns if we are just searching
            # Actually, let's just find columns with "RM "
            col += 1
            if col > ws.max_column * 2: # Fail safe
                break
            continue
            
        # extract RM number
        m = re.search(r"RM\s+(\d+)", str(name), re.I)
        ordem = m.group(1) if m else None
        
        # Look for date in row 1
        d_val = ws.cell(row=1, column=col).value
        if isinstance(d_val, datetime):
            if d_val > max_date:
                max_date = d_val
        elif isinstance(d_val, str):
            try:
                # Basic string parsing if needed
                d_obj = datetime.strptime(d_val.strip(), "%d/%m/%Y")
                d_val = d_obj
                if d_val > max_date:
                    max_date = d_val
            except:
                d_val = max_date # fallback
        else:
            d_val = max_date # fallback if empty
            
        romaneio_groups.append({
            'date': d_val,
            'col_data': col,
            'col_formula': col + 1,
            'name': name,
            'ordem': ordem
        })
        existing_romaneios[ordem] = True
        col += 2 # Skip the formula column
        
    print(f"Maior data encontrada: {max_date}")

    # Add new romaneios
    new_romaneios = [rd for rd in pdf_data.values() if rd["ordem"] not in existing_romaneios]
    print(f"Novos romaneios a inserir: {len(new_romaneios)}")
    
    new_cols_start = ws.max_column + 1
    # Find next available empty pair
    # A simple way is to just append
    c_idx = ws.max_column + 1
    
    for rd in new_romaneios:
        next_date = max_date + timedelta(days=2)
        max_date = next_date # update max date for subsequent ones
        
        # New Romaneio Data Column
        col_data = c_idx
        ws.cell(row=1, column=col_data, value=next_date)
        ws.cell(row=1, column=col_data).number_format = 'dd/mm/yyyy'
        ws.cell(row=2, column=col_data, value=rd['name'])
        
        for sku, qty in rd['itens'].items():
            if sku in product_rows:
                r = product_rows[sku]
                ws.cell(row=r, column=col_data, value=qty)
                
        # New Formula Column
        col_formula = c_idx + 1
        ws.cell(row=2, column=col_formula, value='ESTOQUE INT')
        
        romaneio_groups.append({
            'date': next_date,
            'col_data': col_data,
            'col_formula': col_formula,
            'name': rd['name'],
            'ordem': rd['ordem']
        })
        
        c_idx += 2

    # Now SORT romaneio_groups by date
    romaneio_groups.sort(key=lambda x: x['date'] if isinstance(x['date'], datetime) else datetime(2023,1,1))
    
    # We will rewrite headers and data starting from column C (col 3)
    # But wait, swapping columns is hard with openpyxl (formulas get messed up if we just copy values).
    # Since formulas in ESTOQUE INT are just = PREV_COL - CURRENT_ROMANEIO, we can rewrite the entire grid of Romaneios!
    
    print("Reescrevendo a matriz de romaneios em ordem...")
    # 1. Read all existing data from col_data before we overwrite
    raw_data = {} # (ordem) -> { row: qty }
    for rg in romaneio_groups:
        col = rg['col_data']
        # If it's a new romaneio it might be beyond the bounds, but we already wrote it above!
        # Let's read it
        rg['items'] = {}
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v and isinstance(v, (int, float)) and v != 0:
                rg['items'][r] = v

    # 2. Clear out columns from 3 to current max_column
    for c in range(3, c_idx + 2):
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=c).value = None

    # 3. Rewrite in order
    target_col = 3
    for i, rg in enumerate(romaneio_groups):
        # Header Row 1 - Date
        c_cell = ws.cell(row=1, column=target_col, value=rg['date'])
        c_cell.number_format = 'dd/mm/yyyy'
        
        # Data validation for Date
        dv = DataValidation(type="date", operator="greaterThanOrEqual", formula1="DATE(2020,1,1)", allow_blank=True)
        dv.error = "Por favor, insira uma data válida."
        dv.errorTitle = "Data Inválida"
        dv.add(c_cell)
        ws.add_data_validation(dv)
        
        # Header Row 2
        ws.cell(row=2, column=target_col, value=rg['name'])
        ws.cell(row=2, column=target_col+1, value='ESTOQUE INT')
        
        # Data and formulas
        # The first ESTOQUE INT is calculated against Column B ("ESTOQUE EM TEMPO REAL")
        # All subsquent ones against the previous ESTOQUE INT (target_col - 1)
        prev_estoque_col_letter = 'B' if target_col == 3 else openpyxl.utils.get_column_letter(target_col - 1)
        curr_romaneio_col_letter = openpyxl.utils.get_column_letter(target_col)
        
        for r in range(3, ws.max_row + 1):
            if r in product_rows.values():
                # Write data
                if r in rg['items']:
                    ws.cell(row=r, column=target_col, value=rg['items'][r])
                else:
                    # Clear it
                    pass
                
                # Write formula for ESTOQUE INT
                formula = f"={prev_estoque_col_letter}{r}-{curr_romaneio_col_letter}{r}"
                ws.cell(row=r, column=target_col+1, value=formula)
                
        target_col += 2

    # Save
    out_file = './PEDIDOS TRANSPORTADORAS 2025 (2).xlsx'
    print(f"Salvando planilha em {out_file}...")
    wb.save(out_file)
    print("Pronto!")

if __name__ == '__main__':
    main()
