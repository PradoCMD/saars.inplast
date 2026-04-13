#!/usr/bin/env python3
"""
Migração: Lê PEDIDOS TRANSPORTADORAS 2025 (3).xlsx, limpa, ordena por data,
aplica formatação de calendário e salva como PEDIDOS TRANSPORTADORAS 2025 (AUTOMAÇÃO).xlsx
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import copy
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "PEDIDOS TRANSPORTADORAS 2025 (3).xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "PEDIDOS TRANSPORTADORAS 2025 (AUTOMAÇÃO).xlsx")


def main():
    print("=" * 60)
    print("  MIGRAÇÃO - Planilha (3) → Planilha (AUTOMAÇÃO)")
    print("=" * 60)
    print(f"\nEntrada: {os.path.basename(INPUT_FILE)}")
    print(f"Saída:   {os.path.basename(OUTPUT_FILE)}")

    # Abrir planilha fonte
    print("\n📂 Abrindo planilha fonte...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=False)
    ws = wb['PROGRAMAÇÃO']
    print(f"   Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

    # ── 1. Ler mapa de produtos (coluna A) ──
    product_rows = {}
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str) and " - " in val:
            sku = val.split(" - ")[0].strip()
            product_rows[sku] = r
    print(f"   Produtos encontrados: {len(product_rows)}")

    # ── 2. Extrair todos os romaneios existentes ──
    romaneio_groups = []
    col = 3
    while col <= ws.max_column:
        name = ws.cell(row=2, column=col).value
        
        # Pular colunas vazias ou "ESTOQUE INT"
        if not name or not str(name).strip():
            col += 1
            continue
        if str(name).strip() == 'ESTOQUE INT':
            col += 1
            continue
        if not str(name).startswith("RM "):
            col += 1
            continue

        # Ler data
        d_val = ws.cell(row=1, column=col).value
        if isinstance(d_val, str):
            try:
                d_val = datetime.strptime(d_val.strip(), "%d/%m/%Y")
            except:
                d_val = None

        # Ler itens desta coluna
        items = {}
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and isinstance(v, (int, float)) and v != 0:
                items[r] = v

        romaneio_groups.append({
            'date': d_val,
            'name': str(name).strip(),
            'items': items,
            'original_col': col,
        })

        # Próximo romaneio (pular col ESTOQUE INT)
        col += 2

    print(f"   Romaneios encontrados: {len(romaneio_groups)}")

    # ── 3. Separar romaneios COM e SEM data ──
    com_data = [rg for rg in romaneio_groups if isinstance(rg['date'], datetime)]
    sem_data = [rg for rg in romaneio_groups if not isinstance(rg['date'], datetime)]
    
    print(f"   Com data: {len(com_data)}")
    print(f"   Sem data: {len(sem_data)}")

    # Ordenar os que têm data
    com_data.sort(key=lambda x: x['date'])
    
    # Romaneios sem data ficam no final (ordem original preservada)
    romaneio_groups_sorted = com_data + sem_data

    # ── 4. Limpar toda a grade de romaneios ──
    print("\n🧹 Limpando grade antiga...")
    for c in range(3, ws.max_column + 2):
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            # Limpar formatação antiga
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.border = Border()
            cell.comment = None
    
    # Limpar data validations existentes
    ws.data_validations.dataValidation = []

    # ── 5. Reescrever em ordem ──
    print("📝 Reescrevendo romaneios ordenados...")
    
    # Estilos
    date_font = Font(bold=True, size=11, color="FFFFFF")
    date_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    date_alignment = Alignment(horizontal="center", vertical="center")
    nodate_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    nodate_font = Font(bold=True, size=11, color="000000")
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    est_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # DataValidation de data (uma única para todas as células)
    dv_date = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="DATE(2020,1,1)",
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    dv_date.promptTitle = "📅 Data do Romaneio"
    dv_date.prompt = "Altere a data para reordenar as colunas.\nFormato: dd/mm/aaaa"
    dv_date.errorTitle = "Data Inválida"
    dv_date.error = "Por favor, insira uma data válida no formato dd/mm/aaaa."
    ws.add_data_validation(dv_date)

    target_col = 3
    for i, rg in enumerate(romaneio_groups_sorted):
        # --- Data (linha 1) ---
        c_cell = ws.cell(row=1, column=target_col)
        if rg['date']:
            c_cell.value = rg['date']
            c_cell.number_format = 'DD/MM/YYYY'
            c_cell.font = date_font
            c_cell.fill = date_fill
        else:
            c_cell.value = None
            c_cell.font = nodate_font
            c_cell.fill = nodate_fill
        c_cell.alignment = date_alignment
        c_cell.border = thin_border
        dv_date.add(c_cell)
        
        c_cell.comment = Comment(
            "📅 Altere esta data para reordenar os romaneios.\n"
            "Na próxima execução, as colunas serão\n"
            "reorganizadas pela ordem das datas.",
            "Automação PCP",
            width=280, height=80,
        )

        # Estoque INT header (linha 1, coluna par)
        est_h = ws.cell(row=1, column=target_col + 1)
        est_h.fill = est_fill
        est_h.border = thin_border

        # --- Nome do romaneio (linha 2) ---
        name_cell = ws.cell(row=2, column=target_col, value=rg['name'])
        name_cell.font = header_font
        name_cell.alignment = header_alignment
        name_cell.border = thin_border

        est_cell = ws.cell(row=2, column=target_col + 1, value='ESTOQUE INT')
        est_cell.font = header_font
        est_cell.alignment = header_alignment
        est_cell.border = thin_border

        # --- Dados e fórmulas (linha 3+) ---
        prev_est_letter = 'B' if target_col == 3 else get_column_letter(target_col - 1)
        curr_rm_letter = get_column_letter(target_col)

        for r in range(3, ws.max_row + 1):
            if r in product_rows.values():
                if r in rg['items']:
                    ws.cell(row=r, column=target_col, value=rg['items'][r])
                formula = f"={prev_est_letter}{r}-{curr_rm_letter}{r}"
                ws.cell(row=r, column=target_col + 1, value=formula)

        has_data_marker = "✅" if rg['date'] else "⚠️ SEM DATA"
        date_str = rg['date'].strftime("%d/%m/%Y") if rg['date'] else "---"
        print(f"  {has_data_marker} Col {target_col:>3d}: {date_str:12s} | {rg['name']}")

        target_col += 2

    # ── 6. Salvar ──
    print(f"\n💾 Salvando: {os.path.basename(OUTPUT_FILE)}")
    wb.save(OUTPUT_FILE)
    
    final_cols = target_col - 1
    print(f"\n{'=' * 60}")
    print(f"  ✅ MIGRAÇÃO CONCLUÍDA!")
    print(f"  Romaneios: {len(romaneio_groups_sorted)}")
    print(f"  Produtos:  {len(product_rows)}")
    print(f"  Colunas:   A até {get_column_letter(final_cols)}")
    print(f"  Com data:  {len(com_data)} (ordenados)")
    print(f"  Sem data:  {len(sem_data)} (no final, destacados em amarelo)")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
