import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import os
import glob
import re
from datetime import datetime, timedelta
import pdfplumber

def ct(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()

def parse_num(v):
    if not v:
        return 0.0
    s = ct(v).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_romaneio_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    
    fname = os.path.basename(pdf_path)
    
    ordem_m = re.search(r"ORDEM\s+DE\s+CARGA[:\s]+(\d+)", text, re.I)
    ordem = ordem_m.group(1) if ordem_m else ""
    if not ordem:
        fn_m = re.search(r"ROMANEIO\s+(\d+)", fname, re.I)
        if fn_m:
            ordem = fn_m.group(1)

    pedidos = []
    ped_m = re.search(
        r"N°\s*UNICO.*?Valor\s*Total\s*\n(.*?)(?=MONTANTE)", text, re.S | re.I
    )
    if ped_m:
        for line in ped_m.group(1).strip().split("\n"):
            l = ct(line)
            if not l:
                continue
            m = re.match(r"^(\d+)\s+(\d+)\s+(\d+)\s*-\s*(.+?)\s+([\d.,]+)$", l)
            if m:
                pedidos.append({
                    "cidade": "", 
                    "parceiro": ct(m.group(4))
                })

    itens = {}
    UNITS = r"(?:PC|UN|PT|KG|CX|MT|M2|LT|FD|SC|RL|CJ|JG|BD|GL|TB|PR|DZ|CT|MH|MO|MM|PAR|PÇ|PE|M)"
    itens_m = re.search(
        r"PRODUTO\s+UND\.\s+QTD\s+NEG\.\s+QTD\s+VOL\.\s*\n(.*?)(?=TOTAL\s*:|\Z)",
        text,
        re.S | re.I,
    )
    if itens_m:
        for line in itens_m.group(1).strip().split("\n"):
            l = ct(line)
            if not l:
                continue
            m = re.match(
                rf"^(\d+)\s*-\s*(.+?)\s+({UNITS})\s+([\d.,]+)(?:\s+([\d.,]+))?\s*$",
                l,
                re.I,
            )
            if m:
                sku = ct(m.group(1))
                qty = parse_num(m.group(4))
                itens[sku] = itens.get(sku, 0) + qty

    cidade_parceiro = pedidos[0]["parceiro"] if pedidos else "Desconhecido"
    
    return {
        "ordem": ordem,
        "name": f"RM {ordem} - {cidade_parceiro}",
        "itens": itens,
        "file": fname
    }

def main():
    print("Iniciando processamento (Incluindo arquivos NOTA)...")
    folder = "/Volumes/logistica/Automação Romaneio Logistica x PCP/"
    pdfs = sorted(glob.glob(os.path.join(folder, "ROMANEIO *.pdf")))
    # REMOVIDO PARA PROCESSAR O NOTA: pdfs = [p for p in pdfs if "NOTA" not in os.path.basename(p).upper()]
    
    print(f"Lendo {len(pdfs)} arquivos de PDF...")
    pdf_data = {} # Agrupados por ORDEM
    for p in pdfs:
        try:
            data = parse_romaneio_pdf(p)
            ordem = data["ordem"]
            if ordem:
                if ordem not in pdf_data:
                    pdf_data[ordem] = data
                else:
                    # Merge itens de romaneios diferentes se for a mesma ordem (Romaneio original + Nota)
                    for sku, qty in data["itens"].items():
                        pdf_data[ordem]["itens"][sku] = pdf_data[ordem]["itens"].get(sku, 0) + qty
                    # Preserva o "nome" se o parceiro não fosse "Desconhecido"
                    if "Desconhecido" in pdf_data[ordem]["name"] and "Desconhecido" not in data["name"]:
                        pdf_data[ordem]["name"] = data["name"]
                        
        except Exception as e:
            print(f"Erro ao ler {p}: {e}")

    print("Carregando planilha alvo...")
    wb = openpyxl.load_workbook('./PEDIDOS TRANSPORTADORAS 2025 (1).xlsx', data_only=False)
    ws = wb['PROGRAMAÇÃO']

    # Read products mapping
    product_rows = {}
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str) and " - " in val:
            sku = val.split(" - ")[0].strip()
            product_rows[sku] = r

    existing_romaneios = {}
    romaneio_groups = []
    max_date = datetime(2023, 1, 1)
    
    col = 3
    while col <= ws.max_column:
        name = ws.cell(row=2, column=col).value
        if not name or not str(name).startswith("RM "):
            col += 1
            if col > ws.max_column * 2:
                break
            continue
            
        m = re.search(r"RM\s+(\d+)", str(name), re.I)
        ordem = m.group(1) if m else None
        
        d_val = ws.cell(row=1, column=col).value
        valid_date = False
        if isinstance(d_val, datetime):
            valid_date = True
        elif isinstance(d_val, str):
            try:
                d_val = datetime.strptime(d_val.strip(), "%d/%m/%Y")
                valid_date = True
            except:
                pass
                
        if valid_date and d_val > max_date:
            max_date = d_val
        if not valid_date:
            d_val = max_date
            
        romaneio_groups.append({
            'date': d_val,
            'col_data': col,
            'col_formula': col + 1,
            'name': name,
            'ordem': ordem
        })
        existing_romaneios[ordem] = True
        col += 2
        
    print(f"Maior data encontrada na planilha original: {max_date}")

    # Add new romaneios
    new_romaneios_data = [rd for rd in pdf_data.values() if rd["ordem"] not in existing_romaneios]
    print(f"Total de novos Romaneios (pós cruzamento, juntando notas): {len(new_romaneios_data)}")
    
    c_idx = ws.max_column + 1
    
    for rd in new_romaneios_data:
        next_date = max_date + timedelta(days=2)
        max_date = next_date
        
        col_data = c_idx
        ws.cell(row=1, column=col_data, value=next_date)
        ws.cell(row=2, column=col_data, value=rd['name'])
        
        for sku, qty in rd['itens'].items():
            if sku in product_rows:
                r = product_rows[sku]
                ws.cell(row=r, column=col_data, value=qty)
                
        col_formula = c_idx + 1
        ws.cell(row=2, column=col_formula, value='ESTOQUE INT')
        
        romaneio_groups.append({
            'date': next_date,
            'col_data': col_data,
            'col_formula': col_formula,
            'name': rd['name'],
            'ordem': rd['ordem']
        })
        
        c_idx += 2

    # SORT by date before rewriting columns
    romaneio_groups.sort(key=lambda x: x['date'] if isinstance(x['date'], datetime) else datetime(2023,1,1))
    
    print("Reescrevendo a matriz de romaneios em ordem e refazendo ESTOQUE INT...")
    
    # Extract existing loaded data from sheet
    for rg in romaneio_groups:
        col = rg['col_data']
        rg['items'] = {}
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v and isinstance(v, (int, float)) and v != 0:
                rg['items'][r] = v

    # Clear current grid
    for c in range(3, c_idx + 2):
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=c).value = None

    # Write in order
    target_col = 3
    for i, rg in enumerate(romaneio_groups):
        c_cell = ws.cell(row=1, column=target_col, value=rg['date'])
        c_cell.number_format = 'dd/mm/yyyy'
        
        dv = DataValidation(type="date", operator="greaterThanOrEqual", formula1="DATE(2020,1,1)", allow_blank=True)
        dv.error = "Por favor, insira uma data válida."
        dv.errorTitle = "Data Inválida"
        dv.add(c_cell)
        ws.add_data_validation(dv)
        
        ws.cell(row=2, column=target_col, value=rg['name'])
        ws.cell(row=2, column=target_col+1, value='ESTOQUE INT')
        
        prev_estoque_col_letter = 'B' if target_col == 3 else openpyxl.utils.get_column_letter(target_col - 1)
        curr_romaneio_col_letter = openpyxl.utils.get_column_letter(target_col)
        
        for r in range(3, ws.max_row + 1):
            if r in product_rows.values():
                if r in rg['items']:
                    ws.cell(row=r, column=target_col, value=rg['items'][r])
                formula = f"={prev_estoque_col_letter}{r}-{curr_romaneio_col_letter}{r}"
                ws.cell(row=r, column=target_col+1, value=formula)
                
        target_col += 2

    out_file = './PEDIDOS TRANSPORTADORAS 2025 (1).xlsx'
    print(f"Sobrescrevendo planilha {out_file}...")
    wb.save(out_file)
    print("Pronto! Planilha finalizada.")

if __name__ == '__main__':
    main()
